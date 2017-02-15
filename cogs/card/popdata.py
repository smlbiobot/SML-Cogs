# -*- coding: utf-8 -*-

"""
The MIT License (MIT)

Copyright (c) 2017 SML

Permission is hereby granted, free of charge, to any person obtaining a
copy of this software and associated documentation files (the "Software"),
to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense,
and/or sell copies of the Software, and to permit persons to whom the
Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS
OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
DEALINGS IN THE SOFTWARE.
"""

from openpyxl import load_workbook
from difflib import SequenceMatcher
import json

cardpop_xlsx_path = 'data/cardpop{}.xlsx'
cardpop_json_path = 'data/cardpop{}.json'
cardpop_path = 'data/cardpop.json'
summary_path = 'data/summary.txt'

cardpop_range_min = 16
cardpop_range_max = 24

similarity_threshold = 0.85

data = {}

out = []


def save_json(filename=None, data=None):
    with open(filename, encoding='utf-8', mode='w') as f:
        json.dump(data, f, indent=4, sort_keys=False, separators=(',',' : '))


def process_data():

    prev_cardpop = None

    for id in range(cardpop_range_min, cardpop_range_max):


        wb = load_workbook(cardpop_xlsx_path.format(id))

        ws = wb.worksheets[0]

        cards = []
        players = []
        cardpop = {}
        decks = {}

        out.append("-" * 80)
        out.append("Snapshot #{}".format(id))

        for i, row in enumerate(ws.iter_rows()):
            # first row is card names
            if i==0:
                for j, cell in enumerate(row):
                    if j:
                        cards.append(cell.value)
                        cardpop[cell.value] = {
                            "count": 0,
                            "change": 0
                        }
            # row 2-101 are card data
            # for older worksheets, it’s len - 2
            elif i<101:
                deck = []
                for j, cell in enumerate(row):
                    if j>0 and cell.value == 1:
                        card_name = cards[j-1]
                        deck.append(card_name)
                        cardpop[card_name]["count"] += 1

                deck = sorted(deck)
                player = {
                    "rank": i,
                    "deck": deck
                }

                # Create deck count
                if len(deck):
                    players.append(player)

                    deck_id = ', '.join(deck)

                    # populate unique decks
                    if deck_id not in decks:
                        decks[deck_id] = {
                            "id": deck_id,
                            "deck": deck,
                            "count": 1,
                            "similarity": {}
                            }
                    else:
                        decks[deck_id]["count"] += 1

        # calculate change
        if prev_cardpop is not None:
            for k, v in cardpop.items():
                # verify card exists previously as some cards may be new
                if k in prev_cardpop:
                    v["change"] = v["count"] - prev_cardpop[k]["count"]


        decks = dict(sorted(decks.items(), key = lambda x: -x[1]["count"]))
        cardpop = dict(sorted(cardpop.items(), key = lambda x: -x[1]["count"]))

        # calculate similarity
        for k, deck in decks.items():
            similarity = {}
            for j, deck2 in decks.items():
                if k != j:
                    similarity[j] = SequenceMatcher(a=k, b=j).ratio()
            similarity = dict(sorted(similarity.items(), key = lambda x: -x[1]))
            deck["similarity"] = similarity
        
        prev_cardpop = cardpop

        out.append("Decks:")
        out.append("Deck similarity threshold: {}".format(similarity_threshold))
        for k, deck in decks.items():
            out.append("{:3d}: {}".format(deck["count"], k))
            # output similarity over threshold
            for sk, sv in deck["similarity"].items():
                if sv > similarity_threshold:
                    out.append("        {:3f}: {}".format(sv, sk))

        out.append("Cards:")
        for k, v in cardpop.items():
            out.append("{:3d} ({:3d}): {}".format(v["count"], v["change"], k))

        data[str(id)] = {
            "players": players,
            "cards": sorted(cards),
            "decks": decks,
            "cardpop": cardpop
            }

    save_json(cardpop_path, data)

    with open(summary_path, encoding="utf-8", mode="w") as f:
        f.write('\n'.join(out))

process_data()







