# -*- coding: utf-8 -*-

"""
The MIT License (MIT)

Copyright (c) 2017 SML

Permission is hereby granted, free of charge, to any person obtaining a
copy of this software and associated documentation files (the "Software"),
to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense,
and/or sell copies of the Software, and to permit persons to whom the
Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS
OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
DEALINGS IN THE SOFTWARE.
"""

from openpyxl import load_workbook
from difflib import SequenceMatcher
import json

cardpop_xlsx_path = '../xlsx/cardpop{}.xlsx'
cardpop_json_path = '../cardpop{}.json'
cardpop_path = '../cardpop.json'
summary_path = '../summary.txt'
crdata_path = '../clashroyale.json'

cardpop_range_min = 8
cardpop_range_max = 26

similarity_threshold = 0.85

data = {}

out = []
card_elixir = {}


def load_json(filename=None):
    with open(filename, encoding='utf-8', mode="r") as f:
        json_data = json.load(f)
    return json_data

def save_json(filename=None, data=None):
    with open(filename, encoding='utf-8', mode='w') as f:
        json.dump(data, f, indent=4, sort_keys=False, separators=(',',' : '))


crdata = load_json(crdata_path)

cpid_dict = {}

def cpid2id(cpid=None):
    if cpid is None:
        return ""
    if cpid in cpid_dict:
        return cpid_dict[cpid]
    for card_id, card_v in crdata["Cards"].items():
        cpid_dict[card_v["cpid"]] = card_id
    return cpid_dict[cpid]

def get_card_elixir(id=None):
    """Return card elixir with cpid."""
    # if cpid is None:
    #     return 0
    # if cpid in card_elixir:
    #     return card_elixir[cpid]
    # for card_id, card_v in crdata["Cards"].items():
    #     card_elixir[card_v["cpid"]] = card_v["elixir"]
    # return card_elixir[cpid]

    return crdata["Cards"][id]["elixir"]

def get_deck_elixir(deck):
    """Calculate average elixir in a deck.

    Decks are in CPID so need to find from crdata"""
    # for card in deck:
    elixir = 0
    for card in deck:
        elixir += get_card_elixir(card)
    return elixir/8



def process_data():

    prev_cardpop = None



    for id in range(cardpop_range_min, cardpop_range_max):

        wb = load_workbook(cardpop_xlsx_path.format(id))

        ws = wb.worksheets[0]

        cards = []
        players = []
        cardpop = {}
        decks = {}

        out.append("-" * 80)
        out.append("Snapshot #{}".format(id))

        for i, row in enumerate(ws.iter_rows()):
            # first row is card names
            if i==0:
                for j, cell in enumerate(row):
                    if j:
                        card_id = cpid2id(cell.value)
                        cards.append(card_id)
                        cardpop[card_id] = {
                            "count": 0,
                            "change": 0
                        }
            # row 2-101 are card data
            # for older worksheets, it’s len - 2
            elif i<101:
                deck = []
                for j, cell in enumerate(row):
                    if j>0 and cell.value == 1:
                        card_id = cards[j-1]
                        # card_id = cpid2id(card_name)
                        deck.append(card_id)
                        cardpop[card_id]["count"] += 1

                deck = sorted(deck)
                player = {
                    "rank": i,
                    "deck": deck
                }

                # Create deck count
                if len(deck):
                    players.append(player)

                    deck_id = ', '.join(deck)

                    # populate unique decks
                    if deck_id not in decks:
                        decks[deck_id] = {
                            "id": deck_id,
                            "deck": deck,
                            "count": 1,
                            "elixir": get_deck_elixir(deck),
                            "similarity": {}
                            }
                    else:
                        decks[deck_id]["count"] += 1

        # calculate change
        if prev_cardpop is not None:
            for k, v in cardpop.items():
                # verify card exists previously as some cards may be new
                if k in prev_cardpop:
                    v["change"] = v["count"] - prev_cardpop[k]["count"]


        decks = dict(sorted(decks.items(), key = lambda x: -x[1]["count"]))
        cardpop = dict(sorted(cardpop.items(), key = lambda x: -x[1]["count"]))

        # calculate similarity
        # for k, deck in decks.items():
        #     similarity = {}
        #     for j, deck2 in decks.items():
        #         if k != j:
        #             similarity[j] = SequenceMatcher(a=k, b=j).ratio()
        #     similarity = dict(sorted(similarity.items(), key = lambda x: -x[1]))
        #     deck["similarity"] = similarity

        prev_cardpop = cardpop

        out.append("Decks:")
        # out.append("Deck similarity threshold: {}".format(similarity_threshold))
        for k, deck in decks.items():
            out.append("{:3d}: {}".format(deck["count"], k))
            # output similarity over threshold
            # for sk, sv in deck["similarity"].items():
            #     if sv > similarity_threshold:
            #         out.append("        {:3f}: {}".format(sv, sk))

        out.append("Cards:")
        for k, v in cardpop.items():
            out.append("{:3d} ({:3d}): {}".format(v["count"], v["change"], k))

        data[str(id)] = {
            "players": players,
            "cards": sorted(cards),
            "decks": decks,
            "cardpop": cardpop
            }

    save_json(cardpop_path, data)

    with open(summary_path, encoding="utf-8", mode="w") as f:
        f.write('\n'.join(out))

process_data()







